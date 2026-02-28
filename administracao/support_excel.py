"""
Suporte para exportação de relatórios em Excel com formatação.

Desenvolvido para o Projeto Final de Engenharia.
Autor: Luciano Pereira Soares <lpsoares@insper.edu.br>
Data: 28 de Fevereiro de 2026
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
from collections import defaultdict


def export_relatorio_projetos_xlsx(projetos, edicao=None):
    """
    Exporta relatório de projetos em formato Excel como planilha tabular com dashboard.
    
    Args:
        projetos: QuerySet de Projeto
        edicao: String com período (ex: "2024.1") ou None
    
    Returns:
        BytesIO: Arquivo Excel em bytes
    """
    
    wb = Workbook()
    
    # ===== ABA 1: DASHBOARD / RESUMO =====
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # ===== ABA 2: DADOS DETALHADOS =====
    ws_dados = wb.create_sheet("Projetos", 1)
    
    # ===== CONFIGURAÇÃO DE ESTILOS =====
    estilo_titulo_principal = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    preenchimento_titulo = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    
    estilo_subtitulo = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    preenchimento_subtitulo = PatternFill(start_color="2E5CB8", end_color="2E5CB8", fill_type="solid")
    
    estilo_label = Font(name='Calibri', size=10, bold=True, color="2F5496")
    preenchimento_label = PatternFill(start_color="E7F0F9", end_color="E7F0F9", fill_type="solid")
    
    estilo_valor = Font(name='Calibri', size=10, bold=True, color="0066CC")
    
    estilo_cabecalho = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill(start_color="2E5CB8", end_color="2E5CB8", fill_type="solid")
    
    estilo_dado = Font(name='Calibri', size=9)
    preenchimento_par = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    preenchimento_impar = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== MONTAGEM DO DASHBOARD =====
    _criar_dashboard(ws_resumo, projetos, edicao, estilo_titulo_principal, preenchimento_titulo,
                     estilo_subtitulo, preenchimento_subtitulo, estilo_label, preenchimento_label,
                     estilo_valor, thin_border)
    
    # ===== MONTAGEM DOS DADOS =====
    _criar_planilha_dados(ws_dados, projetos, edicao, estilo_titulo_principal, preenchimento_titulo,
                          estilo_cabecalho, preenchimento_cabecalho, estilo_dado, 
                          preenchimento_par, preenchimento_impar, thin_border)
    
    # Salva em BytesIO
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    
    return arquivo


def _criar_dashboard(ws, projetos, edicao, estilo_titulo, preench_titulo, estilo_subtitulo,
                     preench_subtitulo, estilo_label, preench_label, estilo_valor, border):
    """Cria aba de resumo/dashboard com estatísticas."""
    
    # Define largura das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    
    # ===== CABEÇALHO =====
    ws.merge_cells('A1:D1')
    celula = ws['A1']
    if edicao:
        celula.value = f"RESUMO DE PROJETOS - Período {edicao}"
    else:
        celula.value = "RESUMO DE PROJETOS - Todos os Períodos"
    celula.font = estilo_titulo
    celula.fill = preench_titulo
    celula.alignment = Alignment(horizontal='center', vertical='center')
    celula.border = border
    ws.row_dimensions[1].height = 25
    
    # Data de geração
    ws.merge_cells('A2:D2')
    celula = ws['A2']
    celula.value = f"Relatório gerado em {datetime.now().strftime('%d de %B de %Y às %H:%M')}"
    celula.font = Font(name='Calibri', size=9, italic=True, color="666666")
    celula.alignment = Alignment(horizontal='right', vertical='center')
    
    linha_atual = 4
    
    # ===== ESTATÍSTICAS GERAIS =====
    ws.merge_cells('A4:B4')
    celula = ws['A4']
    celula.value = "ESTATÍSTICAS GERAIS"
    celula.font = estilo_subtitulo
    celula.fill = preench_subtitulo
    celula.alignment = Alignment(horizontal='left', vertical='center')
    celula.border = border
    linha_atual = 5
    
    # Total de projetos
    _adicionar_stat(ws, linha_atual, "Total de Projetos:", len(list(projetos)), 
                    estilo_label, preench_label, estilo_valor, border)
    linha_atual += 1
    
    # Estatísticas por tipo
    tipos_count = defaultdict(int)
    organizacoes = defaultdict(int)
    areas_interesse = defaultdict(int)
    
    projetos_list = list(projetos)
    
    for projeto in projetos_list:
        tipo = "Regular"
        if projeto.proposta:
            if projeto.proposta.intercambio:
                tipo = "Intercâmbio"
            elif projeto.proposta.empreendendo:
                tipo = "Empreendendo"
            elif projeto.proposta.internacional:
                tipo = "Internacional"
        tipos_count[tipo] += 1
        
        if projeto.organizacao:
            organizacoes[projeto.organizacao.nome] += 1
        
        if projeto.proposta:
            for area in projeto.proposta.areadeinteresse_set.all():
                area_nome = area.area.titulo if area.area else area.outras
                areas_interesse[area_nome] += 1
    
    # Adiciona break
    linha_atual += 1
    
    # Seção Projetos por Tipo
    ws.merge_cells(f'A{linha_atual}:B{linha_atual}')
    celula = ws[f'A{linha_atual}']
    celula.value = "PROJETOS POR TIPO"
    celula.font = estilo_subtitulo
    celula.fill = preench_subtitulo
    celula.alignment = Alignment(horizontal='left', vertical='center')
    celula.border = border
    linha_atual += 1
    
    for tipo, count in sorted(tipos_count.items(), key=lambda x: x[1], reverse=True):
        _adicionar_stat(ws, linha_atual, f"  {tipo}:", count, 
                        estilo_label, preench_label, estilo_valor, border)
        linha_atual += 1
    
    # Adiciona break
    linha_atual += 1
    
    # Seção Projetos por Organização (top 10)
    ws.merge_cells(f'A{linha_atual}:B{linha_atual}')
    celula = ws[f'A{linha_atual}']
    celula.value = "TOP 10 ORGANIZAÇÕES"
    celula.font = estilo_subtitulo
    celula.fill = preench_subtitulo
    celula.alignment = Alignment(horizontal='left', vertical='center')
    celula.border = border
    linha_atual += 1
    
    for org, count in sorted(organizacoes.items(), key=lambda x: x[1], reverse=True)[:10]:
        org_nome = org[:25] + "..." if len(org) > 25 else org
        _adicionar_stat(ws, linha_atual, f"  {org_nome}:", count, 
                        estilo_label, preench_label, estilo_valor, border)
        linha_atual += 1
    
    # Adiciona break
    linha_atual += 1
    
    # Seção Áreas de Interesse
    ws.merge_cells(f'A{linha_atual}:B{linha_atual}')
    celula = ws[f'A{linha_atual}']
    celula.value = "ÁREAS DE INTERESSE"
    celula.font = estilo_subtitulo
    celula.fill = preench_subtitulo
    celula.alignment = Alignment(horizontal='left', vertical='center')
    celula.border = border
    linha_atual += 1
    
    for area, count in sorted(areas_interesse.items(), key=lambda x: x[1], reverse=True):
        area_nome = area[:25] + "..." if len(area) > 25 else area
        _adicionar_stat(ws, linha_atual, f"  {area_nome}:", count, 
                        estilo_label, preench_label, estilo_valor, border)
        linha_atual += 1
    
    # Informações da geração
    linha_atual += 2
    ws.merge_cells(f'A{linha_atual}:D{linha_atual}')
    celula = ws[f'A{linha_atual}']
    celula.value = "Use a aba 'Projetos' para ver detalhes completos de cada projeto"
    celula.font = Font(name='Calibri', size=9, italic=True, color="0066CC")
    celula.alignment = Alignment(horizontal='center', vertical='center')


def _adicionar_stat(ws, linha, label, valor, estilo_label, preench_label, estilo_valor, border):
    """Helper para adicionar uma linha de estatística."""
    celula = ws[f'A{linha}']
    celula.value = label
    celula.font = estilo_label
    celula.fill = preench_label
    celula.alignment = Alignment(horizontal='left', vertical='center')
    celula.border = border
    
    celula = ws[f'B{linha}']
    celula.value = valor
    celula.font = estilo_valor
    celula.fill = preench_label
    celula.alignment = Alignment(horizontal='center', vertical='center')
    celula.border = border
    ws.row_dimensions[linha].height = 18


def _criar_planilha_dados(ws, projetos, edicao, estilo_titulo, preench_titulo, estilo_cabecalho,
                          preench_cabecalho, estilo_dado, preench_par, preench_impar, border):
    """Cria planilha detalhada de projetos com formatação e hyperlinks."""
    
    # Define as colunas
    colunas = [
        "Período",
        "Organização",
        "Título do Projeto",
        "Tipo",
        "Título Original",
        "Estudantes",
        "Orientador",
        "Coorientadores",
        "Resumo",
        "Abstract",
        "Palavras-chave",
        "Descrição (Proposta)",
        "Expectativas",
        "Recursos",
        "Áreas de Interesse",
        "Endereço Org",
        "Website Org"
    ]
    
    # Define largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 40
    ws.column_dimensions['M'].width = 40
    ws.column_dimensions['N'].width = 30
    ws.column_dimensions['O'].width = 25
    ws.column_dimensions['P'].width = 30
    ws.column_dimensions['Q'].width = 25
    
    # ===== CABEÇALHO COM PERÍODO =====
    ws.merge_cells('A1:Q1')
    celula_titulo = ws['A1']
    if edicao:
        celula_titulo.value = f"RELATÓRIO DE PROJETOS CAPSTONE - Período {edicao}"
    else:
        celula_titulo.value = "RELATÓRIO DE PROJETOS CAPSTONE - Todos os Períodos"
    celula_titulo.font = estilo_titulo
    celula_titulo.fill = preench_titulo
    celula_titulo.alignment = Alignment(horizontal='center', vertical='center')
    celula_titulo.border = border
    ws.row_dimensions[1].height = 20
    
    # ===== LINHA DE CABEÇALHO =====
    for col_idx, nome_coluna in enumerate(colunas, 1):
        celula = ws.cell(row=2, column=col_idx)
        celula.value = nome_coluna
        celula.font = estilo_cabecalho
        celula.fill = preench_cabecalho
        celula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celula.border = border
    
    ws.row_dimensions[2].height = 25
    ws.freeze_panes = 'A3'
    
    # ===== PREENCHIMENTO DOS DADOS DOS PROJETOS =====
    linha_atual = 3
    projetos_list = list(projetos)
    
    # Cores para formatação condicional por tipo
    cores_tipo = {
        "Intercâmbio": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Empreendendo": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        "Internacional": PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid"),
        "Regular": preench_par
    }
    
    for projeto in projetos_list:
        # Determina cor para linha baseado no tipo
        tipo_projeto = ""
        if projeto.proposta and projeto.proposta.intercambio:
            tipo_projeto = "Intercâmbio"
        elif projeto.proposta and projeto.proposta.empreendendo:
            tipo_projeto = "Empreendendo"
        elif projeto.proposta and projeto.proposta.internacional:
            tipo_projeto = "Internacional"
        else:
            tipo_projeto = "Regular"
        
        preenchimento_linha = cores_tipo.get(tipo_projeto, preench_par)
        
        # Dados do projeto
        periodo = f"{projeto.ano}.{projeto.semestre}"
        organizacao = projeto.organizacao.nome if projeto.organizacao else ""
        
        titulo_display = (projeto.titulo_final 
                         if (projeto.titulo_final and projeto.titulo_final != projeto.proposta.titulo) 
                         else projeto.get_titulo())
        
        titulo_original = (projeto.proposta.titulo 
                          if projeto.proposta and projeto.titulo_final and projeto.titulo_final != projeto.proposta.titulo 
                          else "")
        
        # Estudantes
        alocacoes = projeto.get_alocacoes()
        estudantes = "; ".join([
            f"{a.aluno.user.get_full_name()}" + (f" ({a.aluno.curso2})" if a.aluno.curso2 else "")
            for a in alocacoes
        ]) if alocacoes else ""
        
        # Orientador
        orientador = (projeto.orientador.user.get_full_name() or projeto.orientador.user.username 
                     if projeto.orientador else "")
        
        # Coorientadores
        coorientadores = "; ".join([
            f"{coor.usuario.get_full_name() or coor.usuario.username}" + 
            (f" ({coor.observacao})" if coor.observacao else "")
            for coor in projeto.coorientador_set.all()
        ]) if projeto.coorientador_set.all() else ""
        
        # Textos do projeto
        resumo = projeto.resumo or ""
        abstract = projeto.abstract or ""
        palavras_chave = projeto.palavras_chave or ""
        
        # Dados da proposta
        descricao_proposta = projeto.proposta.descricao if projeto.proposta else ""
        expectativas = projeto.proposta.expectativas if projeto.proposta else ""
        recursos = projeto.proposta.recursos if projeto.proposta else ""
        
        areas = ""
        if projeto.proposta:
            areas_list = [
                (a.area.titulo if a.area else a.outras) 
                for a in projeto.proposta.areadeinteresse_set.all()
            ]
            areas = "; ".join(areas_list) if areas_list else ""
        
        endereco = projeto.organizacao.endereco if projeto.organizacao else ""
        website = projeto.organizacao.website if projeto.organizacao else ""
        
        # Dados a preencher
        dados = [
            periodo,
            organizacao,
            titulo_display,
            tipo_projeto,
            titulo_original,
            estudantes,
            orientador,
            coorientadores,
            resumo,
            abstract,
            palavras_chave,
            descricao_proposta,
            expectativas,
            recursos,
            areas,
            endereco,
            website
        ]
        
        # Preenche cada coluna
        for col_idx, valor in enumerate(dados, 1):
            celula = ws.cell(row=linha_atual, column=col_idx)
            
            # Adiciona hyperlink para website (coluna Q)
            if col_idx == 17 and valor and (valor.startswith('http://') or valor.startswith('https://')):
                celula.hyperlink = valor
                celula.value = valor
                celula.font = Font(name='Calibri', size=9, color="0066CC", underline="single")
            else:
                celula.value = valor
                celula.font = estilo_dado
            
            celula.fill = preenchimento_linha
            celula.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            celula.border = border
        
        ws.row_dimensions[linha_atual].height = 30
        linha_atual += 1
    
    # Se não houver projetos
    if not projetos_list:
        ws.merge_cells('A3:Q3')
        celula = ws['A3']
        celula.value = "Nenhum projeto disponível para o período selecionado."
        celula.font = Font(name='Calibri', size=10, italic=True, color="666666")
        celula.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        celula.alignment = Alignment(horizontal='center', vertical='center')
        celula.border = border
    
    # ===== ATIVA AutoFilter =====
    if projetos_list:
        ws.auto_filter.ref = f'A2:Q{linha_atual - 1}'
